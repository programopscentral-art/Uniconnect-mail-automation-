import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def export_content_to_excel(courses_df: pd.DataFrame, sessions_df: pd.DataFrame, resources_df: pd.DataFrame, mapping_df: pd.DataFrame) -> io.BytesIO:
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        courses_df.to_excel(writer, index=False, sheet_name='SemesterCourse')
        sessions_df.to_excel(writer, index=False, sheet_name='Session')
        resources_df.to_excel(writer, index=False, sheet_name='SessionResource')
        mapping_df.to_excel(writer, index=False, sheet_name='SemesterCourseSession')
        
        workbook = writer.book
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for sheet_name in ['SemesterCourse', 'Session', 'SessionResource', 'SemesterCourseSession']:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                ws.freeze_panes = 'A2'
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)

    output.seek(0)
    return output
