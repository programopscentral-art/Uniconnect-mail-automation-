import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

def export_to_excel(apd_df: pd.DataFrame, breakdown_df: pd.DataFrame, validation_df: pd.DataFrame) -> io.BytesIO:
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        apd_df.to_excel(writer, index=False, sheet_name='APD_2.0_Generated')
        breakdown_df.to_excel(writer, index=False, sheet_name='Slot_Breakdown')
        validation_df.to_excel(writer, index=False, sheet_name='Validation_Report')
        
        # Access the workbook and sheets to apply formatting
        workbook = writer.book
        
        # Apply standard formatting to all sheets
        for sheet_name in ['APD_2.0_Generated', 'Slot_Breakdown', 'Validation_Report']:
            ws = workbook[sheet_name]
            
            # Freeze header
            ws.freeze_panes = 'A2'
            
            # Bold and color headers
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto column width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        # Conditional formatting for Validation_Report
        vs = workbook['Validation_Report']
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        # Find column indices for Status and Buffer
        headers = [cell.value for cell in vs[1]]
        try:
            status_idx = headers.index('Status') + 1
            buffer_idx = headers.index('Buffer Slots') + 1
            
            for row in range(2, vs.max_row + 1):
                status_cell = vs.cell(row=row, column=status_idx)
                if status_cell.value == 'NOT_FEASIBLE':
                    status_cell.fill = red_fill
                elif status_cell.value == 'WARNING':
                    status_cell.fill = yellow_fill
                
                buffer_cell = vs.cell(row=row, column=buffer_idx)
                if buffer_cell.value and isinstance(buffer_cell.value, (int, float)) and buffer_cell.value < 0:
                    buffer_cell.fill = red_fill
        except ValueError:
            pass

    output.seek(0)
    return output
